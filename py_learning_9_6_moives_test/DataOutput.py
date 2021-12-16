import openpyxl
from openpyxl import Workbook


class DataOutput:
    def __init__(self, file_name):
        wb = Workbook()
        ws = wb.create_sheet()
        self.data_collections = []
        wb.save(file_name)

    def stored_data(self, data):
        if data is None:
            return
        self.data_collections.append(data)

    def output_data(self, file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.worksheets[1]
        ws.append(['movie_id', 'movie_title', 'rating_final', 'r_other_final',
                   'r_picture_final', 'r_directory_final', 'r_story_final', 'user_count',
                   'attitude_count', 'total_box_office + total_box_office_unit',
                   'today_box_office + today_box_office_unit',
                   'rank', 'show_days', 'is_release'])
        i = 2
        for data in self.data_collections:
            ws['i'].append(data)
            i += 1
        wb.save(file_name)
