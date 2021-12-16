import openpyxl
from openpyxl import Workbook
import random


def get_random_information(file_name):
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(['姓名', '课程', '成绩'])

    # 中文名字生成
    first = '赵钱孙李'
    middle = '伟昀琛东'
    last = '坤艳志'
    subjects = ['语文', '数学', '英语']
    for i in range(200):
        line = []
        r = random.randint(1, 100)
        name = random.choice(first)
        # 按照一定概率生成只有两个字的中文名字
        if r > 50:
            name = name + random.choice(middle)
        name = name + random.choice(last)
        # 依次生成姓名、课程名称和成绩
        line.append(name)
        line.append(random.choice(subjects))
        line.append(random.randint(0, 100))

        worksheet.append(line)
    workbook.save(file_name)


def get_result(old_file, new_file):
    # 用于存放数据的字典
    result = dict()

    # 打开原始数据
    workbook = openpyxl.load_workbook(old_file)
    worksheet = workbook.worksheets[0]

    # 遍历原始数据
    for row in worksheet.rows:
        if row[0].value == '姓名':
            continue
        # 姓名、课程名称、本次成绩
        name, subject, grade = row[0].value, row[1].value, row[2].value

        # 获取当前姓名对应的课程名称和成绩
        # 如果result字典中不包含，则返回空字典
        t = result.get(name, {})
        # 获取当前学生的课程成绩，若不存在，返回0
        f = t.get(subject, 0)
        # 只保留该学生该课程的最高成绩
        if grade > f:
            t[subject] = grade
            result[name] = t

        workbook1 = Workbook()
        worksheet1 = workbook1.worksheets[0]
        worksheet1.append(['姓名', '课程', '成绩'])

        # 将result字典中的结果写入Excel表格
        for name, t in result.items():
            print(name, t)
            for subject, grade in t.items():
                worksheet1.append([name, subject, grade])

        workbook1.save(new_file)


if __name__ == '__main__':
    r_old_file = r'excel_test.xlsx'
    r_new_file = r'result.xlsx'
    get_random_information(r_old_file)
    get_result(r_old_file, r_new_file)
